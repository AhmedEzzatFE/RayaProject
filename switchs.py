from openpyxl import load_workbook

workbook = load_workbook(filename='Switches_SSH-data.xlsx')
data = workbook['Sheet1']


workbook = load_workbook(filename='switch 1 Ports.xlsx')
port1 = workbook['Sheet1']

workbook = load_workbook(filename='Switch 2 ports .xlsx')
port2 = workbook['Sheet1']

myswitch=open('switch1.txt','a+')
myswitch2=open('switch2.txt','a+')

for row in data.iter_rows(values_only=True,min_row=2):
    if (row[0]=='SW1'):
        myswitch.write('enable' + '\n' + 'configure terminal' + '\n' + 'hostname ' + str(
            row[0]) + '\n' + 'ip domain-name local' + '\n' + 'crypto key generate rsa' + '\n\n'
                       + 'line vty 0 4' + '\n' + 'transport input ssh' + '\n' + 'login local' + '\n' + 'exit' + '\n' + 'username ' + str(
            row[1]) + ' password ' + str(row[2]) + '\n\n')
        x = input('please enter the default gateway ip of SW1 ')
        myswitch.write('ip default-gateway ' + str(x) + '\n' + 'interface VLAN 1' + '\n' + 'ip address ' + str(
            row[5]) + ' 255.255.255.0' + '\n' + 'exit' + '\n\n')

        countInterface=0
        CountPorts=0
        for i in port1.iter_rows(values_only=True, min_row=2):
            if(i[1]=='Trunk'):
                myswitch.write('interface Ethernet'+str(countInterface)+'/'+str(CountPorts)+'\n'+'switchport trunk encapsulation dot1q'+'\n'+'switchport mode trunk'+'\n'+'switchport trunk allowed Vlan ('+str(i[2])+')'+'\n'+'exit'+'\n')
            elif(i[1]=='Access'):
                myswitch.write('vlan '+str(i[2])+'\n'+'show vlan'+'\n'+'exit'+'\n'+'interface Ethernet'+str(countInterface)+'/'+str(CountPorts)+'\n'+'switch mode access'+'\n'+'switchport access vlan '+str(i[2])+'\n'+'exit'+'\n')

            CountPorts+=1
            if(CountPorts==4):
                countInterface+=1
                CountPorts=0
            if(countInterface==4): countInterface=0

    elif(row[0]=='SW2'):
        myswitch2.write('enable' + '\n' + 'configure terminal' + '\n' + 'hostname ' + str(
            row[0]) + '\n' + 'ip domain-name local' + '\n' + 'crypto key generate rsa' + '\n\n'
                       + 'line vty 0 4' + '\n' + 'transport input ssh' + '\n' + 'login local' + '\n' + 'exit' + '\n' + 'username ' + str(
            row[1]) + ' password ' + str(row[2]) + '\n\n')
        x = input('please enter the default gateway ip of SW2 ')
        myswitch2.write('ip default-gateway ' + str(x) + '\n' + 'interface VLAN 1' + '\n' + 'ip address ' + str(
            row[5]) + ' 255.255.255.0' + '\n' + 'exit' + '\n\n')

        countInterface = 0
        CountPorts = 0
        for i in port2.iter_rows(values_only=True, min_row=2):
            if(i[1]=='Trunk'):
                myswitch2.write('interface Ethernet'+str(countInterface)+'/'+str(CountPorts)+'\n'+'switchport trunk encapsulation dot1q'+'\n'+'switchport mode trunk'+'\n'+'switchport trunk allowed Vlan ('+str(i[2])+')'+'\n'+'exit'+'\n')
            elif(i[1]=='Access'):
                myswitch2.write('vlan '+str(i[2])+'\n'+'show vlan'+'\n'+'exit'+'\n'+'interface Ethernet'+str(countInterface)+'/'+str(CountPorts)+'\n'+'switch mode access'+'\n'+'switchport access vlan '+str(i[2])+'\n'+'exit'+'\n')

            CountPorts+=1
            if(CountPorts==4):
                countInterface+=1
                CountPorts=0
            if(countInterface==4): countInterface=0























