

from openpyxl import load_workbook

workbook = load_workbook(filename='Routers_data.xlsx')


sheet = workbook['Sheet1']



set_of_host = set()
for row in sheet.iter_rows(values_only=True,min_row=2):
    set_of_host.add(row[0])




count = 0
for i in set_of_host:
    f = open("demofile" + str(count) + ".txt", "w+")
    f.write('hostname ' + str(i) + '\n' + 'enable secret ' + str(row[3]) + '\n' + 'ip domain-name local' + '\n' + 'crypto key generate rsa 2048' + '\n' + 'ip ssh version 2'
            + '\n' 'username ' + str(row[1]) + ' password ' + str(row[2]) + '\n' + 'line vty 0 4' + '\n' + 'transport input telnet ssh' + '\n' + 'login local' + '\n' + '\n')
    for row in sheet.iter_rows(values_only=True,min_row=2):
        if row[0] == i:

            f.write( 'interface ethernet 0/' + str(row[4]) + '\n' + 'ip address ' + str(row[5]) + ' 255.255.255.0' + '\n' + '\n')
    network_dis = input('please enter network_dis')
    router_dis = input('please enter router_dis')
    f.write( 'ip route ' + str(network_dis) + ' 255.255.255.0 ' + str(router_dis) + '\n' + '\n')
    count +=1











